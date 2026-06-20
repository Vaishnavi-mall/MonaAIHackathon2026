"""
Shift replacement agent — Problem 02.
Customer: Universitätsklinikum des Saarlandes / UKS (Homburg)

Strategy:
  1. Python applies all 6 eligibility rules deterministically from the Excel data.
  2. AI receives only the pre-filtered list and ranks by tie-breakers + drafts the message.

All stateless — no data written to disk.
"""

import io
import re
import logging
from datetime import datetime

from shared.gemini_client import GeminiClient
from shared.utils import clean_json_response
from .prompts import SYSTEM_PROMPT, SHIFT_REPLACEMENT_PROMPT

logger = logging.getLogger(__name__)
_client = GeminiClient()

VALID_URGENCY_LEVELS = {"low", "medium", "high", "critical"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_workbook(file_bytes: bytes):
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl") from exc
    try:
        return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read Excel file: {exc}") from exc


def _col(headers: list, fragment: str) -> int:
    """First column index whose header contains fragment (case-insensitive). -1 if not found."""
    for i, h in enumerate(headers):
        if fragment.lower() in str(h or "").lower():
            return i
    return -1


def _str(val) -> str:
    return str(val).strip() if val is not None else ""


def _parse_date(shift_details: str) -> datetime | None:
    """Extract the first recognisable date from shift_details text."""
    m = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", shift_details)
    if m:
        return datetime.strptime(m.group(1), "%Y-%m-%d")
    m = re.search(r"\b(\d{2})([-/\.])(\d{2})\2(\d{4})\b", shift_details)
    if m:
        d, _, mo, y = m.group(1), m.group(2), m.group(3), m.group(4)
        return datetime.strptime(f"{y}-{mo}-{d}", "%Y-%m-%d")
    return None


# ── Public helpers ─────────────────────────────────────────────────────────────

def extract_employee_info(file_bytes: bytes, staff_name: str) -> dict | None:
    """Look up a staff member by name in the Roster sheet and return their verified details."""
    try:
        wb = _load_workbook(file_bytes)
    except Exception:
        return None
    if "Roster" not in wb.sheetnames:
        return None

    ws = wb["Roster"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    headers = [_str(h) for h in rows[0]]
    first_i = _col(headers, "first")
    last_i  = _col(headers, "last")
    role_i  = _col(headers, "role")
    dept_i  = _col(headers, "department")
    cert_i  = _col(headers, "certif")
    stat_i  = _col(headers, "status")
    needle  = staff_name.strip().lower()

    for row in rows[1:]:
        first = _str(row[first_i]) if first_i >= 0 else ""
        last  = _str(row[last_i])  if last_i  >= 0 else ""
        full  = f"{first} {last}".strip().lower()
        if not full:
            continue
        if full == needle or needle in full or full in needle:
            return {
                "name":           f"{first} {last}".strip(),
                "role":           _str(row[role_i]) if role_i >= 0 else "",
                "department":     _str(row[dept_i]) if dept_i >= 0 else "",
                "certifications": _str(row[cert_i]) if cert_i >= 0 else "",
                "status":         _str(row[stat_i]) if stat_i >= 0 else "",
            }
    return None


# ── Core eligibility filter ────────────────────────────────────────────────────

def _filter_eligible(wb, absent_info: dict, shift_date: datetime) -> list[dict]:
    """
    Apply all 6 eligibility rules from the scenario deterministically.

    Rules:
      1. Role: Registered Nurse or Charge Nurse (if absent employee is RN)
      2. Certifications: must hold all certs the absent employee has
      3. Status: Active
      4. OFF on the shift date (code == 'O' in Weekly_Schedule)
      5. Adequately rested: not currently on shift; last clock-out not after 08:30 today
      6. Hours cap: scheduled_hrs (next 7d) + 12 ≤ Max Hrs/Week
    """
    # ── Roster ────────────────────────────────────────────────────────────────
    r_ws   = wb["Roster"]
    r_rows = list(r_ws.iter_rows(values_only=True))
    r_hdrs = [_str(h) for h in r_rows[0]]

    ri_id       = _col(r_hdrs, "employee id")
    ri_first    = _col(r_hdrs, "first")
    ri_last     = _col(r_hdrs, "last")
    ri_role     = _col(r_hdrs, "role")
    ri_dept     = _col(r_hdrs, "department")
    ri_cert     = _col(r_hdrs, "certif")
    ri_contract = _col(r_hdrs, "contract")
    ri_max_hrs  = _col(r_hdrs, "max hrs")
    ri_overtime = _col(r_hdrs, "overtime")
    ri_status   = _col(r_hdrs, "status")
    ri_persona  = _col(r_hdrs, "persona")
    ri_last_out = _col(r_hdrs, "last clock out")
    ri_phone    = _col(r_hdrs, "phone")

    # ── Weekly_Schedule ───────────────────────────────────────────────────────
    s_ws   = wb["Weekly_Schedule"]
    s_rows = list(s_ws.iter_rows(values_only=True))
    s_hdrs = [_str(h) for h in s_rows[0]]

    # Column for the target date (e.g. "06/20")
    target_mm_dd = shift_date.strftime("%m/%d")
    date_col     = next((i for i, h in enumerate(s_hdrs) if target_mm_dd in h), -1)

    # All date columns (for manual hours calculation — formula values may be missing)
    date_pat  = re.compile(r"\d{1,2}/\d{2}")
    all_date_cols = [i for i, h in enumerate(s_hdrs) if date_pat.search(h)]
    # "next 7 days" = skip the first date column (yesterday)
    next7_cols = all_date_cols[1:]

    s_name_col = _col(s_hdrs, "name")

    # Build lookup: name_lower → {shift_code, scheduled_hrs}
    schedule = {}
    for row in s_rows[1:]:
        name = _str(row[s_name_col]) if s_name_col >= 0 else ""
        if not name:
            continue
        code = _str(row[date_col]) if date_col >= 0 else ""
        hrs  = sum(12 for i in next7_cols if _str(row[i]).upper() in ("D", "N"))
        schedule[name.lower()] = {"code": code, "hrs": hrs}

    # ── Eligibility criteria ──────────────────────────────────────────────────
    absent_role  = absent_info.get("role", "").lower()
    absent_certs = {c.strip().upper() for c in absent_info.get("certifications", "").split(",") if c.strip()}
    absent_name  = absent_info.get("name", "").strip().lower()

    if "registered nurse" in absent_role:
        valid_roles = {"registered nurse", "charge nurse"}
    elif "charge nurse" in absent_role:
        valid_roles = {"charge nurse"}
    else:
        valid_roles = {absent_role}

    today = datetime.now()
    eligible = []

    for row in r_rows[1:]:
        first = _str(row[ri_first]) if ri_first >= 0 else ""
        last  = _str(row[ri_last])  if ri_last  >= 0 else ""
        name  = f"{first} {last}".strip()
        if not name or name.lower() == absent_name:
            continue

        role     = _str(row[ri_role])     if ri_role     >= 0 else ""
        certs_s  = _str(row[ri_cert])     if ri_cert     >= 0 else ""
        status   = _str(row[ri_status])   if ri_status   >= 0 else ""
        last_out = row[ri_last_out]       if ri_last_out >= 0 else None

        try:
            max_hrs = float(row[ri_max_hrs] or 0) if ri_max_hrs >= 0 else 0.0
        except (TypeError, ValueError):
            max_hrs = 0.0

        sched    = schedule.get(name.lower(), {})
        code     = sched.get("code", "")
        sched_h  = sched.get("hrs", 0)

        # Rule 1: role
        if role.lower() not in valid_roles:
            continue
        # Rule 2: certifications
        certs = {c.strip().upper() for c in certs_s.split(",") if c.strip()}
        if absent_certs and not absent_certs.issubset(certs):
            continue
        # Rule 3: status
        if status.lower() != "active":
            continue
        # Rule 4: OFF on target date
        if date_col < 0 or code != "O":
            continue
        # Rule 5: adequately rested
        lo_str = _str(last_out)
        if "on shift" in lo_str.lower():
            continue  # currently working
        if isinstance(last_out, datetime):
            if last_out.date() == today.date() and (last_out.hour, last_out.minute) >= (8, 30):
                continue  # finishing a Day shift today
        # Rule 6: hours cap
        if max_hrs > 0 and (sched_h + 12) > max_hrs:
            continue

        contract    = _str(row[ri_contract]) if ri_contract >= 0 else ""
        overtime_ok = _str(row[ri_overtime]).lower() if ri_overtime >= 0 else ""
        persona     = _str(row[ri_persona])  if ri_persona  >= 0 else ""
        phone       = _str(row[ri_phone])    if ri_phone    >= 0 else ""
        dept        = _str(row[ri_dept])     if ri_dept     >= 0 else ""
        emp_id      = _str(row[ri_id])       if ri_id       >= 0 else ""

        eligible.append({
            "name":                   name,
            "employee_id":            emp_id,
            "role":                   role,
            "department":             dept,
            "certifications":         certs_s,
            "contract":               contract,
            "max_hrs_per_week":       int(max_hrs),
            "scheduled_hrs_next_7d":  sched_h,
            "hours_headroom":         int(max_hrs - sched_h),
            "overtime_ok":            overtime_ok == "yes",
            "persona":                persona,
            "phone":                  phone,
        })

    return eligible


# ── Main entry point ───────────────────────────────────────────────────────────

def process_shift_request(file_bytes: bytes, absent_staff: str, shift_details: str) -> dict:
    """
    Find shift replacements.

    Deterministically filters eligible candidates using all 6 scenario rules,
    then calls the AI only for tie-breaking ranking and message drafting.
    """
    if not absent_staff or not absent_staff.strip():
        raise ValueError("absent_staff cannot be empty")
    if not shift_details or not shift_details.strip():
        raise ValueError("shift_details cannot be empty")

    wb = _load_workbook(file_bytes)

    name_only    = absent_staff.split(",")[0].strip()
    absent_info  = extract_employee_info(file_bytes, name_only) or {
        "name": name_only, "role": "", "department": "", "certifications": "", "status": "Active",
    }
    shift_date = _parse_date(shift_details)

    # ── Deterministic filtering ───────────────────────────────────────────────
    eligible = []
    if shift_date and "Roster" in wb.sheetnames and "Weekly_Schedule" in wb.sheetnames:
        eligible = _filter_eligible(wb, absent_info, shift_date)

    logger.info(
        "Shift replacement — absent=%s shift=%s eligible_count=%d",
        absent_staff, shift_details, len(eligible),
    )

    # ── Build prompt ──────────────────────────────────────────────────────────
    absent_block = (
        f"ABSENT EMPLOYEE (verified directly from Roster — treat as ground truth):\n"
        f"  Name:           {absent_info['name']}\n"
        f"  Role:           {absent_info.get('role', 'Unknown')}\n"
        f"  Department:     {absent_info.get('department', 'Unknown')}\n"
        f"  Certifications: {absent_info.get('certifications', 'Unknown')}\n"
    )

    if eligible:
        lines = []
        for i, c in enumerate(eligible, 1):
            lines.append(
                f"{i}. {c['name']} ({c['employee_id']}) — {c['role']}, {c['department']}\n"
                f"   Contract: {c['contract']} | Overtime OK: {'Yes' if c['overtime_ok'] else 'No'}\n"
                f"   Hours headroom: {c['hours_headroom']} hrs (scheduled {c['scheduled_hrs_next_7d']} / max {c['max_hrs_per_week']})\n"
                f"   Certs: {c['certifications']}\n"
                f"   Persona: {c['persona']}\n"
                f"   Phone: {c['phone']}"
            )
        candidates_block = (
            f"ELIGIBLE CANDIDATES — {len(eligible)} found (all 6 eligibility rules already applied):\n"
            + "\n".join(lines)
            + "\n\nRank these candidates by the tie-breaker criteria and return the JSON."
        )
    else:
        candidates_block = (
            "No candidates passed all 6 eligibility rules. "
            "Return an empty available_staff list and explain in the draft_message."
        )

    prompt = SHIFT_REPLACEMENT_PROMPT.format(
        absent_block=absent_block,
        candidates_block=candidates_block,
        absent_staff=absent_staff,
        shift_details=shift_details,
    )

    raw    = _client.generate(prompt, system_prompt=SYSTEM_PROMPT)
    result = clean_json_response(raw)

    # ── Override with deterministic ground-truth values ───────────────────────
    if absent_info.get("role"):
        result["required_role"] = absent_info["role"]
    if shift_date:
        result["shift_date"] = shift_date.strftime("%Y-%m-%d")
    if result.get("urgency_level") not in VALID_URGENCY_LEVELS:
        result["urgency_level"] = "high"
    result["human_review_required"] = True

    # Attach phone numbers from the deterministic candidate list
    phone_map = {c["name"].lower(): c["phone"] for c in eligible}
    for staff in result.get("available_staff", []):
        key = staff.get("name", "").lower()
        if key in phone_map and phone_map[key]:
            staff["phone"] = phone_map[key]

    logger.info(
        "Replacement found — urgency=%s candidates=%d",
        result.get("urgency_level"),
        len(result.get("available_staff", [])),
    )

    return result


# ── Backwards-compatible shims (used by server.py) ────────────────────────────

def parse_schedule_file(file_bytes: bytes) -> str:
    """Convert Excel to plain text. Kept for API compatibility."""
    wb = _load_workbook(file_bytes)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [_str(c) for c in row]
            if any(c for c in cells):
                lines.append("\t".join(cells))
    return "\n".join(lines)


def find_shift_replacements(
    schedule_data: str,
    absent_staff: str,
    shift_details: str,
    absent_employee_info: dict | None = None,
) -> dict:
    """Deprecated shim — server.py now calls process_shift_request directly."""
    raise NotImplementedError(
        "Use process_shift_request(file_bytes, absent_staff, shift_details) instead."
    )
